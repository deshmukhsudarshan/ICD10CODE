import openpyxl
from openpyxl import load_workbook
import nlpaug.augmenter.word as naw

def augment_text(description, num_examples=5):
    """
    Generate new examples for an ICD-10 code description using NLP augmentation.
    :param description: Original description to base new examples on.
    :param num_examples: Number of new examples to generate.
    :return: List of new descriptions.
    """
    # Create a synonym-based augmenter
    augmenter = naw.SynonymAug(aug_src='wordnet')

    augmented_examples = []
    for _ in range(num_examples):
        try:
            # Generate an augmented version of the text
            augmented_text = augmenter.augment(description)
            if augmented_text:
                augmented_examples.append(augmented_text)
        except Exception as e:
            print(f"Error during augmentation: {e}")
    
    return augmented_examples

def process_excel(input_file, output_file):
    """
    Read an Excel file, generate new examples for each ICD-10 description,
    and save the updated file.
    :param input_file: Path to the input Excel file.
    :param output_file: Path to save the output Excel file.
    """
    try:
        # Load workbook and select the first sheet
        wb = load_workbook(input_file)
        ws = wb.active
        
        # Assume the first column has ICD-10 codes and the second has descriptions
        for row in ws.iter_rows(min_row=2, max_col=2):
            icd_code = row[0].value
            description = row[1].value
            
            if description:
                print(f"Generating examples for ICD-10 code: {icd_code}")
                new_examples = augment_text(description)
                for example in new_examples:
                    ws.append([icd_code, example])
        
        # Save the modified workbook
        wb.save(output_file)
        print(f"New examples added and saved to {output_file}.")
    
    except Exception as e:
        print(f"Error processing Excel file: {e}")

# Input and Output file paths
input_file = "icd10_descriptions.xlsx"  # Replace with your file path
output_file = "icd10_descriptions_with_examples.xlsx"

process_excel(input_file, output_file)
